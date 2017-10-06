#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Thu Oct  5 23:33:03 2017

@author: Lesile
"""

"""
###Chapter12 Excel集成(Excel Integration)
Python与Excel集成可以起到以下作用:
    操纵工具:使用Python可以与Excel电子表格交互并加以操纵
    数据处理器:Python可以向电子表格提供数据,从电子表格处理数据
    分析引擎:Python可以向电子表格提供完整的分析能力，称为VBA编程的成熟替代品
"""

"""
###12.1 基本电子表格交互(Basic Spreadsheet Interaction)
本小结主要说明生成、读取和操纵Excel电子表格文件的方法
"""
import numpy as np
import pandas as pd
import xlrd, xlwt
import xlsxwriter
path = 'E:\\Github\\ReadingNotes\\PythonForFinance\\'

###12.1.1 生成工作簿 Generating Workbooks (xls)
#生成工作簿
wb = xlwt.Workbook()
wb
#在工作簿中加入一个或多个工作表
wb.add_sheet('first_sheet', cell_overwrite_ok=True)
#查看工作表的索引号
wb.get_active_sheet()
#为工作表定义一个别名
ws_1 = wb.get_sheet(0)
ws_1
#实例化和别名可以合并为一步
ws_2 = wb.add_sheet('second_sheet')

#生成模拟数据
data = np.arange(1, 65).reshape((8, 8))
data
#在表格中写入一个数字
ws_1.write(0, 0, 100)# write 100 in cell "A1"
#样本数据可以批量写入到工作簿中
for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws_1.write(r,c,float(data[c,r]))
        ws_2.write(r,c,float(data[r,c]))
#如果不转换为float类型将会出现以下错误:
    #Exception: Unexpected data type <class 'numpy.int32'>
    
#保存表格(Workbook类的save方法可以将整个工作簿对象保存到磁盘中)
wb.save(path + 'workbook.xls')

###12.1.2  生成工作簿 Generating Workbooks (xslx)
#创建一个工作簿对象
wb = xlsxwriter.Workbook(path + 'workbook.xlsx')
#创建工作表对象
ws_1 = wb.add_worksheet('first_sheet')
ws_2 = wb.add_worksheet('second_sheet')
#将数据写入工作表对象
for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws_1.write(r, c, data[c, r])
        ws_2.write(r, c, data[r, c])
#关闭工作簿对象
wb.close()


##xlsx还有许多生成工作簿对象的选项（添加图像）
wb = xlsxwriter.Workbook(path + 'chart.xlsx')
ws = wb.add_worksheet()

# write cumsum of random values in first column
values = np.random.standard_normal(15).cumsum()
ws.write_column('A1', values)

# create a new chart object
chart = wb.add_chart({'type': 'line'})

# add a series to the chart
chart.add_series({'values': '=Sheet1!$A$1:$A$15',
                  'marker': {'type': 'diamond'},})
  # series with markers (here: diamond)

# insert the chart
ws.insert_chart('C1', chart)
wb.close()

###12.1.3  从工作簿中读取 Reading from Workbooks
#xlwt的姊妹库xlrd负责从电子表格文件(工作簿)中读取数据
book = xlrd.open_workbook(path + 'workbook.xlsx')
book
#sheet_names方法提供特定工作簿对象中所有工作表对象名称
book.sheet_names()
#工作表可以通过名称或索引访问
sheet_1 = book.sheet_by_name('first_sheet')
sheet_2 = book.sheet_by_index(1)
sheet_1
sheet_2.name
#工作表对象的重要属性是ncols和nrows，分别表示包含数据的列数和行数
sheet_1.ncols, sheet_1.nrows
#通过cell访问单元格
cl = sheet_1.cell(0, 0)
cl.value
#单元格的属性
cl.ctype

sheet_2.row(3)
sheet_2.col(3)
sheet_1.col_values(3, start_rowx=3, end_rowx=7)
sheet_1.row_values(3, start_colx=3, end_colx=7)
#读取工作簿中的所有数据
for c in range(sheet_1.ncols):
    for r in range(sheet_1.nrows):
        print ('%i' % sheet_1.cell(r, c).value, end=' ')

###12.1.4  Using OpenPyxl
import openpyxl as oxl
#生成工作簿对象
wb = oxl.Workbook()
#创建一个工作表对象
ws = wb.create_sheet(index=0, title='oxl_sheet')
#将数据写入工作表
for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws.cell(row=r+1, column=c+1).value = data[c, r]
        # creates a Cell object and assigns a value
#关闭文件对象
wb.save(path + 'oxl_book.xlsx')

#利用openpyxl读取工作簿
wb = oxl.load_workbook(path + 'oxl_book.xlsx')
ws = wb.get_active_sheet()
cell = ws['B4']
cell.column
cell.row
cell.value
ws['B1':'B4']
for cell in ws['B1':'B4']:
    print(cell[0].value)

	
for row in ws['B1':'B4']:
    for cell in row:
        print(cell.value, end=' ')
    print
    

###12.1.5  Using pandas for Reading and Writing
#pandas读取excel
df_1 = pd.read_excel(path + 'workbook.xlsx',
                     'first_sheet', header=None)
df_2 = pd.read_excel(path + 'workbook.xlsx',
                     'second_sheet', header=None)
#添加列名称
import string
columns = []
for c in range(data.shape[0]):
    columns.append(string.ascii_uppercase[c])
columns
df_1.columns = columns
df_2.columns = columns
df_1
df_2
#利用pandas写出Excel到磁盘
df_1.to_excel(path + 'new_book_1.xlsx', 'my_sheet')

#利用xlrd读取
wbn = xlrd.open_workbook(path + 'new_book_1.xlsx')
wbn.sheet_names()

#现在要将多个DataFrame对象写入单个电子表格文件，需要一个ExcelWriter对象
wbw = pd.ExcelWriter(path + 'new_book_2.xlsx')
df_1.to_excel(wbw, 'first_sheet')
df_2.to_excel(wbw, 'second_sheet')
wbw.save()

#检查确认
wbn = xlrd.open_workbook(path + 'new_book_2.xlsx')
wbn.sheet_names()

#更大的数据的读写
data = np.random.rand(20, 10000)
data.nbytes
df = pd.DataFrame(data)
%time df.to_excel(path + 'data.xlsx', 'data_sheet')
%time np.save(path + 'data', data)

%time df = pd.read_excel(path + 'data.xlsx', 'data_sheet')
%time  data = np.load(path + 'data.npy')

"""
###12.2 用Python编写Excel脚本
向Excel电子表格输出Python的分析能力(这是技术要求更高的工作).
方法一:Python库PyXLL通过提供Excel加载项(add-in,Microsoft增强Excel功能的技术)输出Python的手段
方法二:DataNitro公司提供一个解决方案，可以将Python和Excel完全集成，使Python称为VBA的替代品
不过以上两种解决方案都是需要许可证的商用商品
"""
#由于DataNitro是商业付费软件，暂不学习

"""
###12.3 xlwings
基本功能:可结合 VBA 实现对 Excel 编程，强大的数据输入分析能力，同时拥有丰富的接口，
        结合 pandas/numpy/matplotlib 轻松应对 Excel 数据处理工作。
和DataNitro相比,xlwings是开源免费的。
xlwings
"""
import xlwings as xw
#连接到excel
filename = path + 'new_book_2.xlsx'
workbook = xw.Book(filename)#连接excel文件
#连接到指定单元格
data_range = workbook.sheets('first_sheet').range('A1')
#写入数据
data_range.value = [1111,2222,3333]
#保存
workbook.save()

"""
DataNitro和xlwings方法使我们可以将Excel当做一个灵活和强大的通用GUI,并将其与Python的分析能力相结合,真可谓两全其美.
"""



